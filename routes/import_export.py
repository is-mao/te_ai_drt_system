from config import Config
from models import db
from models.defect_report import DefectReport
from flask import Blueprint, request, jsonify, render_template, send_file
from routes.auth import login_required
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import_export_bp = Blueprint("import_export", __name__, url_prefix="")

# Column mapping: Excel header -> model field
COLUMN_MAPPING = {
    # Cesium / raw data headers
    "Record Time": "record_time",
    "Serial Number": "sn",
    "Failing Test Name": "failure",
    "Machine": "server",
    "UUT Type": "pcap_n",
    "Test Area": "station",
    # Export / template headers
    "Record Time (UTC)": "record_time",
    "SN": "sn",
    "Failure": "failure",
    "Server": "server",
    "PCAP/N": "pcap_n",
    "Station": "station",
    "Week#": "week_number",
    # Common headers
    "Defect class": "defect_class",
    "Defect value": "defect_value",
    "Root cause": "root_cause",
    "Action": "action",
    "PN": "pn",
    "Component SN": "component_sn",
    "LOG": "log_content",
}

# Export column order
EXPORT_COLUMNS = [
    ("bu", "BU"),
    ("week_number", "Week#"),
    ("pcap_n", "PCAP/N"),
    ("station", "Station"),
    ("server", "Server"),
    ("sn", "SN"),
    ("record_time", "Record Time (UTC)"),
    ("failure", "Failure"),
    ("defect_class", "Defect class"),
    ("defect_value", "Defect value"),
    ("root_cause", "Root cause"),
    ("action", "Action"),
    ("pn", "PN"),
    ("component_sn", "Component SN"),
]

LOG_COLUMNS = [
    ("log_content", "LOG"),
    ("sequence_log", "Sequence Log"),
    ("buffer_log", "Buffer Log"),
]


@import_export_bp.route("/import", methods=["GET"])
@login_required
def import_page():
    return render_template("import.html", bu_options=Config.BU_OPTIONS, defect_classes=Config.DEFECT_CLASSES)


@import_export_bp.route("/pending", methods=["GET"])
@login_required
def pending_page():
    return render_template("pending.html", bu_options=Config.BU_OPTIONS)


@import_export_bp.route("/api/import/excel", methods=["POST"])
@login_required
def import_excel():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    file = request.files["file"]
    bu = request.form.get("bu", "").strip().upper()

    if not bu or bu not in Config.BU_OPTIONS:
        return jsonify({"success": False, "error": f'Invalid BU. Must be one of: {", ".join(Config.BU_OPTIONS)}'}), 400

    if not file.filename or not file.filename.endswith(".xlsx"):
        return jsonify({"success": False, "error": "Only .xlsx files are accepted"}), 400

    try:
        file_bytes = BytesIO(file.read())
        # Try openpyxl first, fall back to zipfile+XML if styles parsing fails
        try:
            file_bytes.seek(0)
            wb = openpyxl.load_workbook(file_bytes, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
        except Exception:
            file_bytes.seek(0)
            headers, all_rows = _read_xlsx_raw(file_bytes)
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to read Excel file: {str(e)}"}), 400

    if not headers or all(h is None for h in headers):
        return jsonify({"success": False, "error": "Excel file has no header row"}), 400

    # Map column indices to model fields (case-insensitive)
    # Handle duplicate headers: first 'SN' -> sn, second 'SN' -> component_sn
    col_map_lower = {k.lower(): v for k, v in COLUMN_MAPPING.items()}
    col_map = {}
    seen_fields = set()
    for idx, header in enumerate(headers):
        if header is None:
            continue
        header_str = str(header).strip()
        field = col_map_lower.get(header_str.lower())
        if field:
            if field in seen_fields:
                # Duplicate: second 'SN' maps to component_sn
                if field == "sn":
                    col_map[idx] = "component_sn"
                    continue
                else:
                    continue  # skip other duplicates
            col_map[idx] = field
            seen_fields.add(field)

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(all_rows, start=2):
        try:
            row_data = {}
            for col_idx, value in enumerate(row):
                if col_idx in col_map:
                    field = col_map[col_idx]
                    row_data[field] = value

            # Skip empty rows
            if not any(row_data.values()):
                continue

            # Parse record_time
            record_time = row_data.get("record_time")
            if record_time:
                if isinstance(record_time, datetime):
                    pass  # already a datetime
                elif isinstance(record_time, str):
                    # Try as Excel serial number first (raw XML fallback returns numbers as strings)
                    try:
                        serial = float(record_time)
                        from datetime import timedelta

                        excel_epoch = datetime(1899, 12, 30)
                        record_time = excel_epoch + timedelta(days=serial)
                    except (ValueError, TypeError):
                        for fmt in (
                            "%Y-%m-%d %H:%M:%S.%f",
                            "%Y-%m-%d %H:%M:%S",
                            "%Y/%m/%d %H:%M:%S",
                            "%m/%d/%Y %H:%M:%S",
                            "%Y-%m-%d",
                            "%m/%d/%Y",
                        ):
                            try:
                                record_time = datetime.strptime(record_time.strip(), fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            record_time = None
                else:
                    record_time = None
                row_data["record_time"] = record_time

            # Auto-calculate week_number from record_time
            rt = row_data.get("record_time")
            if isinstance(rt, datetime):
                iso_week = rt.isocalendar()[1]
                year_short = rt.year % 100
                row_data["week_number"] = f"{year_short}WK{iso_week:02d}"

            # Duplicate check by (sn + record_time)
            sn = row_data.get("sn")
            rt = row_data.get("record_time")
            if sn and rt:
                sn_str = str(sn).strip()
                existing = DefectReport.query.filter_by(sn=sn_str, record_time=rt).first()
                if existing:
                    skipped += 1
                    continue

            # Convert all values to strings where appropriate
            for field in (
                "week_number",
                "pcap_n",
                "station",
                "server",
                "sn",
                "failure",
                "defect_class",
                "defect_value",
                "root_cause",
                "action",
                "pn",
                "component_sn",
                "log_content",
            ):
                val = row_data.get(field)
                if val is not None:
                    row_data[field] = str(val).strip()

            report = DefectReport(
                bu=bu,
                week_number=row_data.get("week_number"),
                pcap_n=row_data.get("pcap_n"),
                station=row_data.get("station"),
                server=row_data.get("server"),
                sn=row_data.get("sn"),
                record_time=row_data.get("record_time"),
                failure=row_data.get("failure"),
                defect_class=row_data.get("defect_class"),
                defect_value=row_data.get("defect_value"),
                root_cause=row_data.get("root_cause"),
                action=row_data.get("action"),
                pn=row_data.get("pn"),
                component_sn=row_data.get("component_sn"),
                log_content=row_data.get("log_content"),
            )
            db.session.add(report)
            imported += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return (
            jsonify(
                {
                    "success": False,
                    "error": f"Database error: {str(e)}",
                    "imported": 0,
                    "skipped": skipped,
                    "errors": errors,
                }
            ),
            500,
        )

    return jsonify({"success": True, "imported": imported, "skipped": skipped, "errors": errors})


@import_export_bp.route("/api/export/excel", methods=["GET"])
@login_required
def export_excel():
    bu = request.args.get("bu")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    station = request.args.get("station")
    defect_class = request.args.get("defect_class")
    defect_value = request.args.get("defect_value")
    search = request.args.get("search")

    query = DefectReport.query.filter(db.or_(DefectReport.status == "complete", DefectReport.status.is_(None)))

    if bu and bu.upper() in Config.BU_OPTIONS:
        query = query.filter(DefectReport.bu == bu.upper())
    if station:
        query = query.filter(DefectReport.station.ilike(f"%{station}%"))
    if defect_class:
        query = query.filter(DefectReport.defect_class == defect_class)
    if defect_value:
        query = query.filter(DefectReport.defect_value == defect_value)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            db.or_(
                DefectReport.sn.ilike(search_term),
                DefectReport.failure.ilike(search_term),
                DefectReport.defect_value.ilike(search_term),
                DefectReport.root_cause.ilike(search_term),
            )
        )
    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(DefectReport.record_time >= dt_from)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d")
            dt_to = dt_to.replace(hour=23, minute=59, second=59)
            query = query.filter(DefectReport.record_time <= dt_to)
        except ValueError:
            pass

    records = query.order_by(DefectReport.record_time.desc()).all()

    include_log = request.args.get("exclude_log") != "1"
    columns = EXPORT_COLUMNS + LOG_COLUMNS if include_log else EXPORT_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defect Reports"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Write headers
    for col_idx, (field, header_name) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        for col_idx, (field, _) in enumerate(columns, start=1):
            value = getattr(record, field, "")
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Set column widths
    col_widths = {
        "BU": 8,
        "Week#": 10,
        "PCAP/N": 15,
        "Station": 18,
        "Server": 18,
        "SN": 22,
        "Record Time (UTC)": 22,
        "Failure": 30,
        "Defect class": 16,
        "Defect value": 20,
        "Root cause": 30,
        "Action": 30,
        "PN": 18,
        "Component SN": 22,
        "LOG": 40,
        "Sequence Log": 40,
        "Buffer Log": 40,
    }
    for col_idx, (field, header_name) in enumerate(columns, start=1):
        width = col_widths.get(header_name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"defect_reports_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@import_export_bp.route("/api/export/template", methods=["GET"])
@login_required
def export_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    template_headers = [
        "Week#",
        "PCAP/N",
        "Station",
        "Server",
        "SN",
        "Record Time (UTC)",
        "Failure",
        "Defect class",
        "Defect value",
        "Root cause",
        "Action",
        "PN",
        "Component SN",
        "LOG",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header_name in enumerate(template_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="drt_import_template.xlsx",
    )


# ---------------------------------------------------------------------------
# Cesium Raw Data Import
# ---------------------------------------------------------------------------

# Cesium column mapping: Excel header -> model field
CESIUM_COLUMN_MAPPING = {
    "Record Time (UTC)": "record_time",
    "Serial Number": "sn",
    "Failing Test Name": "failure",
    "Machine": "server",
    "UUT Type": "pcap_n",
    "Test Area": "station",
}


def _calc_week_number(dt):
    """Calculate week number: Jan 1 = W01, each 7 days = 1 week. Format: 26WK01."""
    if not dt:
        return ""
    if isinstance(dt, datetime):
        day_of_year = dt.timetuple().tm_yday
        year_short = dt.year % 100
    else:
        day_of_year = dt.timetuple().tm_yday
        year_short = dt.year % 100
    week_num = (day_of_year - 1) // 7 + 1
    return f"{year_short}WK{week_num:02d}"


def _read_xlsx_raw(file_bytes):
    """Read xlsx using zipfile+XML, bypassing openpyxl style parsing entirely."""
    import zipfile
    import xml.etree.ElementTree as ET

    zf = zipfile.ZipFile(file_bytes)

    # Read shared strings (text values are stored here)
    shared_strings = []
    if "xl/sharedStrings.xml" in zf.namelist():
        ss_xml = zf.read("xl/sharedStrings.xml")
        ss_root = ET.fromstring(ss_xml)
        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        for si in ss_root.findall(".//s:si", ns):
            texts = si.findall(".//s:t", ns)
            shared_strings.append("".join(t.text or "" for t in texts))

    # Read first sheet
    sheet_xml = zf.read("xl/worksheets/sheet1.xml")
    sheet_root = ET.fromstring(sheet_xml)
    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    rows_data = []
    for row_el in sheet_root.findall(".//s:sheetData/s:row", ns):
        row_cells = {}
        for cell in row_el.findall("s:c", ns):
            ref = cell.get("r", "")  # e.g. "A1", "B1"
            cell_type = cell.get("t", "")
            val_el = cell.find("s:v", ns)
            value = val_el.text if val_el is not None else None

            if value is not None:
                if cell_type == "s":  # shared string
                    idx = int(value)
                    value = shared_strings[idx] if idx < len(shared_strings) else ""
                elif cell_type == "b":  # boolean
                    value = bool(int(value))

            # Extract column letter from ref
            col_letter = "".join(c for c in ref if c.isalpha())
            row_cells[col_letter] = value

        rows_data.append(row_cells)

    zf.close()

    if not rows_data:
        return [], []

    # First row is headers
    header_row = rows_data[0]
    # Get all column letters in order
    all_cols = sorted(set().union(*(r.keys() for r in rows_data)), key=lambda x: (len(x), x))
    headers = [header_row.get(c) for c in all_cols]
    data_rows = []
    for row in rows_data[1:]:
        data_rows.append([row.get(c) for c in all_cols])

    return headers, data_rows


@import_export_bp.route("/api/import/cesium", methods=["POST"])
@login_required
def import_cesium():
    """Import Cesium raw data as draft records."""
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    file = request.files["file"]
    bu = request.form.get("bu", "").strip().upper()

    if not bu or bu not in Config.BU_OPTIONS:
        return jsonify({"success": False, "error": f'Invalid BU. Must be one of: {", ".join(Config.BU_OPTIONS)}'}), 400

    if not file.filename or not file.filename.endswith(".xlsx"):
        return jsonify({"success": False, "error": "Only .xlsx files are accepted"}), 400

    try:
        file_bytes = BytesIO(file.read())
        # Try openpyxl first, fall back to zipfile+XML if styles parsing fails
        try:
            file_bytes.seek(0)
            wb = openpyxl.load_workbook(file_bytes, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
        except Exception:
            file_bytes.seek(0)
            headers, data_rows = _read_xlsx_raw(file_bytes)
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to read Excel file: {str(e)}"}), 400

    if not headers or not data_rows:
        return jsonify({"success": False, "error": "Excel file has no data rows"}), 400

    # Map column indices to model fields
    col_map = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        header_str = str(header).strip()
        if header_str in CESIUM_COLUMN_MAPPING:
            col_map[idx] = CESIUM_COLUMN_MAPPING[header_str]

    if not col_map:
        return jsonify({"success": False, "error": "No matching Cesium columns found in the file"}), 400

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(data_rows, start=2):
        try:
            row_data = {}
            for col_idx, field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and val != "":
                    row_data[field] = val

            # Skip empty rows
            if not any(row_data.values()):
                continue

            # Parse record_time
            record_time = row_data.get("record_time")
            if record_time is not None:
                if isinstance(record_time, datetime):
                    pass
                elif isinstance(record_time, str):
                    # Try as Excel serial number first (raw XML returns numbers as strings)
                    try:
                        serial = float(record_time)
                        from datetime import timedelta

                        excel_epoch = datetime(1899, 12, 30)
                        record_time = excel_epoch + timedelta(days=serial)
                    except (ValueError, TypeError):
                        # Try date string formats
                        for fmt in (
                            "%Y-%m-%d %H:%M:%S.%f",
                            "%Y-%m-%d %H:%M:%S",
                            "%Y/%m/%d %H:%M:%S",
                            "%m/%d/%Y %H:%M:%S",
                            "%Y-%m-%d",
                            "%m/%d/%Y",
                        ):
                            try:
                                record_time = datetime.strptime(record_time.strip(), fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            record_time = None
                else:
                    record_time = None
                row_data["record_time"] = record_time

            # Convert to strings
            for field in ("sn", "failure", "station", "pcap_n", "server"):
                val = row_data.get(field)
                if val is not None:
                    row_data[field] = str(val).strip()

            # Duplicate check by (sn + record_time)
            sn = row_data.get("sn")
            rt = row_data.get("record_time")
            if sn and rt:
                existing = DefectReport.query.filter_by(sn=sn, record_time=rt).first()
                if existing:
                    skipped += 1
                    continue

            # Auto-calculate week number
            week_number = _calc_week_number(rt) if rt else ""

            report = DefectReport(
                bu=bu,
                sn=row_data.get("sn"),
                record_time=rt,
                failure=row_data.get("failure"),
                station=row_data.get("station"),
                pcap_n=row_data.get("pcap_n"),
                server=row_data.get("server"),
                week_number=week_number,
                status="draft",
            )
            db.session.add(report)
            imported += 1

        except Exception as e:
            errors.append(f"Row {row_idx + 2}: {str(e)}")

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return (
            jsonify(
                {
                    "success": False,
                    "error": f"Database error: {str(e)}",
                    "imported": 0,
                    "skipped": skipped,
                    "errors": errors,
                }
            ),
            500,
        )

    return jsonify({"success": True, "imported": imported, "skipped": skipped, "errors": errors})


# ---------------------------------------------------------------------------
# Draft Records Management API
# ---------------------------------------------------------------------------


@import_export_bp.route("/api/draft-records", methods=["GET"])
@login_required
def api_draft_records():
    """Get draft records for the import page."""
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 25, type=int)
    bu = request.args.get("bu", "").strip()

    query = DefectReport.query.filter_by(status="draft")
    if bu:
        query = query.filter(DefectReport.bu == bu)

    query = query.order_by(DefectReport.created_at.desc())
    total = query.count()
    from math import ceil

    total_pages = ceil(total / per_page) if per_page else 1
    records = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify(
        {
            "data": [r.to_dict(include_log=False) for r in records],
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
        }
    )


@import_export_bp.route("/api/draft-records/<int:id>/complete", methods=["POST"])
@login_required
def api_draft_complete(id):
    """Mark a draft record as complete (after filling in required fields)."""
    record = DefectReport.query.get_or_404(id)
    if record.status != "draft":
        return jsonify({"success": False, "error": "Record is not a draft"}), 400

    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "No data provided"}), 400

    # Update fields from form
    updatable = [
        "bu",
        "week_number",
        "pcap_n",
        "station",
        "server",
        "sn",
        "failure",
        "defect_class",
        "defect_value",
        "root_cause",
        "action",
        "pn",
        "component_sn",
        "log_content",
        "sequence_log",
        "buffer_log",
    ]
    for field in updatable:
        if field in data:
            setattr(record, field, data[field])

    if "record_time" in data and data["record_time"]:
        for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                record.record_time = datetime.strptime(data["record_time"], fmt)
                break
            except ValueError:
                continue

    record.status = "complete"

    # Validate required fields before completing
    missing = []
    if not record.bu:
        missing.append("BU")
    if not record.sn:
        missing.append("SN")
    if not record.station:
        missing.append("Station")
    if not record.failure:
        missing.append("Failure")
    if missing:
        return jsonify({"success": False, "error": f"Missing required fields: {', '.join(missing)}"}), 400

    db.session.commit()

    return jsonify({"success": True, "data": record.to_dict(include_log=True)})


@import_export_bp.route("/api/draft-records/<int:id>", methods=["DELETE"])
@login_required
def api_draft_delete(id):
    """Delete a draft record."""
    record = DefectReport.query.get_or_404(id)
    if record.status != "draft":
        return jsonify({"success": False, "error": "Record is not a draft"}), 400
    db.session.delete(record)
    db.session.commit()
    return jsonify({"success": True})


@import_export_bp.route("/api/draft-records/batch-delete", methods=["POST"])
@login_required
def api_draft_batch_delete():
    """Batch delete draft records by IDs."""
    data = request.get_json()
    if not data or not isinstance(data.get("ids"), list) or not data["ids"]:
        return jsonify({"success": False, "error": "No IDs provided"}), 400

    ids = [int(i) for i in data["ids"] if str(i).isdigit()]
    if not ids:
        return jsonify({"success": False, "error": "Invalid IDs"}), 400

    deleted = DefectReport.query.filter(DefectReport.id.in_(ids), DefectReport.status == "draft").delete(
        synchronize_session=False
    )
    db.session.commit()

    return jsonify({"success": True, "deleted": deleted})
